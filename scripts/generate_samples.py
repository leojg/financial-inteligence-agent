"""Generate synthetic bank statement sample data for finance-intelligence-agent.

Usage:
    python generate_samples.py                          # Jan 2026 (default)
    python generate_samples.py --start 2026-03-01 --end 2026-03-31
    python generate_samples.py --start 2026-06-01 --end 2026-06-30 --out ./data/june
"""

import argparse
import json
import logging
import textwrap
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

logger = logging.getLogger(__name__)

DEFAULT_OUT = "./data"
BASELINE_START = date(2026, 1, 1)
BASELINE_END   = date(2026, 1, 31)

# ── Shared transaction data ──────────────────────────────────────────────────
# Categories align with reconciliator categorize prompt: Withdrawals (ATM/cash),
# Subscriptions (SaaS / recurring), plus Groceries, Dining, Transport, Utilities,
# Healthcare, Entertainment, Shopping, Travel, Education, Salary, Freelance,
# Transfer, Fees & Charges, Rent, Other Income, Other.

# Itaú - Cuenta Corriente (UYU) — XLSX
ITAU_TRANSACTIONS = [
    ("2026-01-02", "SUPERMERCADO DISCO",         -2840.00, "UYU", "Groceries"),
    ("2026-01-03", "ANTEL TELEFONIA",             -890.00,  "UYU", "Utilities"),
    ("2026-01-05", "FARMACIA DEL PUEBLO",         -420.00,  "UYU", "Healthcare"),
    ("2026-01-06", "SUPERMERCADO TIENDA INGLESA", -3120.00, "UYU", "Groceries"),
    ("2026-01-07", "YPF COMBUSTIBLES",            -1800.00, "UYU", "Transport"),
    ("2026-01-08", "NETFLIX",                     -720.00,  "UYU", "Subscriptions"),
    ("2026-01-09", "UBER",                        -380.00,  "UYU", "Transport"),
    ("2026-01-10", "RESTAURANT LA PASCUALA",      -1560.00, "UYU", "Dining"),
    ("2026-01-12", "SUPERMERCADO DISCO",          -1950.00, "UYU", "Groceries"),
    ("2026-01-13", "OSE AGUA",                    -310.00,  "UYU", "Utilities"),
    ("2026-01-14", "UTE ELECTRICIDAD",            -1240.00, "UYU", "Utilities"),
    ("2026-01-15", "SALARIO ENERO",               52000.00, "UYU", "Salary"),
    ("2026-01-15", "FARMACIA DEL PUEBLO",         -580.00,  "UYU", "Healthcare"),
    ("2026-01-16", "GIMNASIO SMART FIT",          -990.00,  "UYU", "Healthcare"),
    ("2026-01-17", "SUPERMERCADO TIENDA INGLESA", -2340.00, "UYU", "Groceries"),
    ("2026-01-18", "RAPPI DELIVERY",              -640.00,  "UYU", "Dining"),
    ("2026-01-19", "YPF COMBUSTIBLES",            -1800.00, "UYU", "Transport"),
    ("2026-01-20", "LIBRERIA PAPELERIA DON",      -450.00,  "UYU", "Shopping"),
    ("2026-01-21", "RESTAURANT EL PALENQUE",      -2100.00, "UYU", "Dining"),
    ("2026-01-22", "TRANSFER TO BROU",            -8000.00, "UYU", "Transfer"),
    ("2026-01-23", "SUPERMERCADO DISCO",          -1780.00, "UYU", "Groceries"),
    ("2026-01-24", "SPOTIFY",                     -360.00,  "UYU", "Subscriptions"),
    ("2026-01-25", "UBER",                        -420.00,  "UYU", "Transport"),
    ("2026-01-26", "FARMACIA URUGUAYA",           -890.00,  "UYU", "Healthcare"),
    ("2026-01-27", "SUPERMERCADO TIENDA INGLESA", -2680.00, "UYU", "Groceries"),
    ("2026-01-28", "CINE LIFE ALFAVILLE",         -780.00,  "UYU", "Entertainment"),
    ("2026-01-29", "RESTAURANT LA CIGALE",        -3200.00, "UYU", "Dining"),
    ("2026-01-30", "YPF COMBUSTIBLES",            -1800.00, "UYU", "Transport"),
    ("2026-01-31", "RAPPI DELIVERY",              -520.00,  "UYU", "Dining"),
]

# BROU - Caja de Ahorros (UYU) — XLSX
BROU_TRANSACTIONS = [
    ("2026-01-01", "SALDO ANTERIOR",              15000.00, "UYU", "Other Income"),
    ("2026-01-03", "PAGO ALQUILER",               -18500.00,"UYU", "Rent"),
    ("2026-01-05", "TRANSFER FROM ITAU",           8000.00, "UYU", "Transfer"),
    ("2026-01-06", "SUPERMERCADO DISCO",          -1950.00, "UYU", "Groceries"),
    ("2026-01-08", "ABITAB PAGO SERVICIOS",       -1200.00, "UYU", "Utilities"),
    ("2026-01-10", "RETIRO CAJERO ATM",           -3000.00, "UYU", "Withdrawals"),
    ("2026-01-12", "COBRO FREELANCE WEB",          9500.00, "UYU", "Freelance"),
    ("2026-01-14", "FARMACIA DEL PUEBLO",          -420.00, "UYU", "Healthcare"),
    ("2026-01-15", "PAGO TARJETA CREDITO",        -6500.00, "UYU", "Fees & Charges"),
    ("2026-01-17", "RETIRO CAJERO ATM",           -2000.00, "UYU", "Withdrawals"),
    ("2026-01-19", "YPF COMBUSTIBLES",            -1800.00, "UYU", "Transport"),
    ("2026-01-20", "ANTEL FIBRA OPTICA",           -650.00, "UYU", "Utilities"),
    ("2026-01-21", "SUPERMERCADO DISCO",          -2100.00, "UYU", "Groceries"),
    ("2026-01-22", "COBRO FREELANCE APP",          7200.00, "UYU", "Freelance"),
    ("2026-01-24", "RETIRO CAJERO ATM",           -2500.00, "UYU", "Withdrawals"),
    ("2026-01-25", "RESTAURANT LA PASCUALA",      -1560.00, "UYU", "Dining"),
    ("2026-01-27", "ABITAB PAGO DGI",             -3200.00, "UYU", "Fees & Charges"),
    ("2026-01-28", "SUPERMERCADO TIENDA INGLESA", -1890.00, "UYU", "Groceries"),
    ("2026-01-29", "RETIRO CAJERO ATM",           -2000.00, "UYU", "Withdrawals"),
    ("2026-01-30", "PAGO MUTUAL MEDICA",          -1100.00, "UYU", "Healthcare"),
    ("2026-01-31", "SALDO FINAL",                    0.00,  "UYU", "Other"),
]

# Wise - USD Account (USD) — PDF
WISE_TRANSACTIONS = [
    ("2026-01-02", "GITHUB COPILOT",               -10.00,  "USD", "Subscriptions"),
    ("2026-01-03", "DIGITAL OCEAN DROPLET",        -24.00,  "USD", "Subscriptions"),
    ("2026-01-05", "REMOTE JOB PAYMENT JAN",      1800.00,  "USD", "Salary"),
    ("2026-01-06", "ADOBE CREATIVE CLOUD",         -54.99,  "USD", "Subscriptions"),
    ("2026-01-07", "NAMECHEAP DOMAIN",             -15.88,  "USD", "Subscriptions"),
    ("2026-01-08", "CHATGPT PLUS",                 -20.00,  "USD", "Subscriptions"),
    ("2026-01-10", "AMAZON WEB SERVICES",          -38.42,  "USD", "Subscriptions"),
    ("2026-01-12", "UDEMY COURSE PURCHASE",        -29.99,  "USD", "Education"),
    ("2026-01-14", "FIGMA PROFESSIONAL",           -15.00,  "USD", "Subscriptions"),
    ("2026-01-15", "TRANSFER TO LOCAL BANK",      -500.00,  "USD", "Transfer"),
    ("2026-01-16", "NOTION TEAM PLAN",             -16.00,  "USD", "Subscriptions"),
    ("2026-01-17", "DIGITAL OCEAN DROPLET",        -24.00,  "USD", "Subscriptions"),
    ("2026-01-18", "UPWORK FREELANCE INCOME",      620.00,  "USD", "Freelance"),
    ("2026-01-20", "GITHUB COPILOT",               -10.00,  "USD", "Subscriptions"),
    ("2026-01-21", "ANTHROPIC API USAGE",          -42.80,  "USD", "Subscriptions"),
    ("2026-01-22", "ZOOM SUBSCRIPTION",            -15.99,  "USD", "Subscriptions"),
    ("2026-01-23", "AMAZON WEB SERVICES",          -41.17,  "USD", "Subscriptions"),
    ("2026-01-24", "1PASSWORD FAMILY",              -4.99,  "USD", "Subscriptions"),
    ("2026-01-25", "UPWORK FREELANCE INCOME",      480.00,  "USD", "Freelance"),
    ("2026-01-27", "ANTHROPIC API USAGE",         -189.40,  "USD", "Subscriptions"),
    ("2026-01-28", "TAILSCALE VPN",                -18.00,  "USD", "Subscriptions"),
    ("2026-01-30", "DIGITAL OCEAN DROPLET",        -24.00,  "USD", "Subscriptions"),
    ("2026-01-31", "GITHUB COPILOT",               -10.00,  "USD", "Subscriptions"),
]

# VISA Credit Card — PDF
VISA_TRANSACTIONS = [
    ("2026-01-03", "SUPERMERCADO DISCO",          -3200.00, "UYU", "Groceries"),
    ("2026-01-04", "RESTAURANT PANORAMICO",       -4800.00, "UYU", "Dining"),
    ("2026-01-05", "ZARA PUNTA CARRETAS",         -5600.00, "UYU", "Shopping"),
    ("2026-01-06", "FARMACIA DEL PUEBLO",          -420.00, "UYU", "Healthcare"),
    ("2026-01-07", "SPOTIFY",                      -360.00, "UYU", "Subscriptions"),
    ("2026-01-09", "APPLE STORE",                -12800.00, "UYU", "Shopping"),
    ("2026-01-10", "UBER",                         -380.00, "UYU", "Transport"),
    ("2026-01-11", "SUPERMERCADO TIENDA INGLESA", -2890.00, "UYU", "Groceries"),
    ("2026-01-13", "RESTAURANT DON PEPERONE",     -1980.00, "UYU", "Dining"),
    ("2026-01-14", "DECATHLON MONTEVIDEO",        -3400.00, "UYU", "Shopping"),
    ("2026-01-15", "NETFLIX",                      -720.00, "UYU", "Subscriptions"),
    ("2026-01-16", "HOTEL COTTAGE COLONIA",       -8900.00, "UYU", "Travel"),
    ("2026-01-17", "YPF COMBUSTIBLES",            -1800.00, "UYU", "Transport"),
    ("2026-01-18", "SUPERMERCADO DISCO",          -2640.00, "UYU", "Groceries"),
    ("2026-01-19", "LIBRERIAS EL CLUB",            -780.00, "UYU", "Shopping"),
    ("2026-01-20", "RAPPI DELIVERY",               -640.00, "UYU", "Dining"),
    ("2026-01-21", "CINES LIFE ALFAVILLE",         -780.00, "UYU", "Entertainment"),
    ("2026-01-22", "GIMNASIO SMART FIT",           -990.00, "UYU", "Healthcare"),
    ("2026-01-23", "RESTAURANT LA TABLA",         -2200.00, "UYU", "Dining"),
    ("2026-01-24", "FARMACIA URUGUAYA",            -890.00, "UYU", "Healthcare"),
    ("2026-01-25", "IKEA ONLINE",                 -6700.00, "UYU", "Shopping"),
    ("2026-01-26", "UBER",                         -460.00, "UYU", "Transport"),
    ("2026-01-28", "SUPERMERCADO TIENDA INGLESA", -3100.00, "UYU", "Groceries"),
    ("2026-01-29", "ANTEL TELEFONIA",              -890.00, "UYU", "Utilities"),
    ("2026-01-30", "RESTAURANT LA CIGALE",        -3200.00, "UYU", "Dining"),
    ("2026-01-31", "PAGO MINIMO VISA",            -5000.00, "UYU", "Fees & Charges"),
]


def _eval_cross_bank_bait_rows() -> tuple[list[tuple], list[tuple]]:
    """Itaú vs BROU rows: same date + amount, different merchants (non-dup eval bait)."""
    itau: list[tuple] = []
    brou: list[tuple] = []
    for i in range(15):
        day = 14 + i
        d = f"2026-01-{day:02d}"
        amt = -248.0 - float(i) * 3.0
        itau.append((d, f"EVAL BAIT SHOP {i:02d}", amt, "UYU", "Shopping"))
        brou.append((d, f"EVAL BAIT OTHER {i:02d}", amt, "UYU", "Dining"))
    return itau, brou


# Suspicious-activity eval (Itaú): outlier dining, rapid identical charges, round transfer
SUSPICIOUS_ITAU_ROWS = [
    ("2026-01-04", "RESTAURANT LA PARRILLA",     -18500.00, "UYU", "Dining"),
    ("2026-01-04", "EVAL DINE BASELINE 1",       -1500.00,  "UYU", "Dining"),
    ("2026-01-04", "EVAL DINE BASELINE 2",       -1600.00,  "UYU", "Dining"),
    ("2026-01-18", "TRANSFERENCIA EVAL ROUND",   -50000.00, "UYU", "Transfer"),
] + [("2026-01-20", "TIENDA XYZ EVAL RAPID", -800.00, "UYU", "Shopping")] * 5


_bait_itau, _bait_brou = _eval_cross_bank_bait_rows()
ITAU_TRANSACTIONS = ITAU_TRANSACTIONS + SUSPICIOUS_ITAU_ROWS + [
    ("2026-01-11", "EVAL ALIAS NETFLIX",     -720.00, "UYU", "Subscriptions"),
    ("2026-01-12", "EVAL UBER EATS STAR",    -450.00, "UYU", "Dining"),
    ("2026-01-13", "EVAL MERCADOLIBRE",      -800.00, "UYU", "Shopping"),
] + _bait_itau

BROU_TRANSACTIONS = BROU_TRANSACTIONS + _bait_brou

VISA_TRANSACTIONS = VISA_TRANSACTIONS + [
    ("2026-01-11", "EVAL NF STAR NETFLIX",   -720.00, "UYU", "Subscriptions"),
    ("2026-01-12", "EVAL UBER EATS DN",      -450.00, "UYU", "Dining"),
    ("2026-01-13", "EVAL MELI MERCADOLIBRE", -792.00, "UYU", "Shopping"),
]


EVAL_ALIAS_DUPLICATE_SPECS: list[dict] = [
    {
        "account_a": "Itaú Corriente",
        "merchant_a": "EVAL ALIAS NETFLIX",
        "account_b": "VISA Credit",
        "merchant_b": "EVAL NF STAR NETFLIX",
        "amount":     -720.0,
    },
    {
        "account_a": "Itaú Corriente",
        "merchant_a": "EVAL UBER EATS STAR",
        "account_b": "VISA Credit",
        "merchant_b": "EVAL UBER EATS DN",
        "amount":     -450.0,
    },
]

EVAL_FUZZY_DUPLICATE_SPECS: list[dict] = [
    {
        "account_a": "Itaú Corriente",
        "merchant_a": "EVAL MERCADOLIBRE",
        "amount_a":   -800.0,
        "account_b": "VISA Credit",
        "merchant_b": "EVAL MELI MERCADOLIBRE",
        "amount_b":   -792.0,
    },
]

TARGET_NON_DUPLICATE_PAIR_COUNT = 30


# ── Date-range utilities ─────────────────────────────────────────────────────

def shift_transactions(transactions, start: date, end: date):
    """Re-map transaction dates from the Jan-2026 baseline to [start, end]."""
    baseline_span = (BASELINE_END - BASELINE_START).days  # 30
    new_span      = (end - start).days

    result = []
    for d, m, a, c, cat in transactions:
        orig = date.fromisoformat(d)
        offset_days = (orig - BASELINE_START).days
        scaled = round(offset_days * new_span / baseline_span)
        new_d = start + timedelta(days=min(scaled, new_span))
        result.append((new_d.isoformat(), m, a, c, cat))
    return result


def period_label(start: date, end: date) -> str:
    return f"{start.strftime('%d/%m/%Y')} – {end.strftime('%d/%m/%Y')}"


def month_label(start: date) -> str:
    return start.strftime("%B %Y").upper()


# ── XLSX generators ──────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="003366")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CREDIT_FONT  = Font(color="1F7A1F", name="Arial", size=10)
DEBIT_FONT   = Font(color="CC0000", name="Arial", size=10)
ROW_FILL_ALT = PatternFill("solid", start_color="EEF4FF")
THIN         = Side(style="thin", color="CCCCCC")
BORDER       = Border(bottom=Border(bottom=THIN).bottom)


def _write_xlsx(path, sheet_name, bank_header_lines, columns, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for i, line in enumerate(bank_header_lines, 1):
        ws.cell(row=i, column=1, value=line).font = Font(bold=(i == 1), name="Arial", size=11 if i == 1 else 9)
    ws.merge_cells("A1:F1")

    header_row = len(bank_header_lines) + 2

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, header_row + 1):
        alt = r_idx % 2 == 0
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            if alt:
                cell.fill = ROW_FILL_ALT
            if c_idx == 3:
                cell.font = CREDIT_FONT if (val or 0) >= 0 else DEBIT_FONT
                cell.number_format = '#,##0.00'

    total_row = header_row + len(rows) + 1
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True, name="Arial", size=10)
    amt_col = get_column_letter(3)
    ws.cell(total_row, 3, f"=SUM({amt_col}{header_row+1}:{amt_col}{header_row+len(rows)})").font = Font(bold=True, name="Arial", size=10)
    ws.cell(total_row, 3).number_format = '#,##0.00'

    widths = [14, 36, 14, 10, 24, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)


def generate_itau_xlsx(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    path = f"{out}/itau_cuenta_corriente_{month}.xlsx"
    _write_xlsx(
        path,
        sheet_name="Cuenta Corriente",
        bank_header_lines=[
            "Banco Itaú Uruguay S.A.",
            "Cuenta Corriente N° ****-4821",
            "Titular: JUAN CARLOS RODRIGUEZ",
            f"Período: {period_label(start, end)}",
        ],
        columns=["Fecha", "Descripción", "Importe (UYU)", "Moneda", "Cuenta", "Referencia"],
        rows=[
            (d, m, a, c, "Itaú Cuenta Corriente ****4821", f"REF{1000+i:04d}")
            for i, (d, m, a, c, _) in enumerate(transactions)
        ],
    )
    logger.info("Generated %s", path)


def generate_brou_xlsx(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    path = f"{out}/brou_caja_ahorros_{month}.xlsx"
    _write_xlsx(
        path,
        sheet_name="Caja de Ahorros",
        bank_header_lines=[
            "Banco de la República Oriental del Uruguay (BROU)",
            "Caja de Ahorros N° ****-7203",
            "Titular: JUAN CARLOS RODRIGUEZ",
            f"Período: {period_label(start, end)}",
        ],
        columns=["Fecha", "Descripción", "Importe (UYU)", "Moneda", "Cuenta", "Comprobante"],
        rows=[
            (d, m, a, c, "BROU Caja de Ahorros ****7203", f"BROU{2000+i:04d}")
            for i, (d, m, a, c, _) in enumerate(transactions)
        ],
    )
    logger.info("Generated %s", path)


# ── PDF generators ───────────────────────────────────────────────────────────

def _build_pdf(path, title, subtitle, account_info, transactions, currency_note):
    doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    title_style = ParagraphStyle("BankTitle", fontSize=16, fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#003366"), spaceAfter=4)
    sub_style   = ParagraphStyle("BankSub",   fontSize=10, fontName="Helvetica",
                                  textColor=colors.HexColor("#555555"), spaceAfter=2)
    note_style  = ParagraphStyle("Note",      fontSize=8,  fontName="Helvetica-Oblique",
                                  textColor=colors.grey, spaceAfter=12)

    story = []
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(subtitle, sub_style))
    for line in account_info:
        story.append(Paragraph(line, sub_style))
    story.append(Paragraph(currency_note, note_style))
    story.append(Spacer(1, 0.3*cm))

    header = ["Date", "Description", "Amount", "Currency", "Account"]
    table_data = [header]
    for d, m, a, c, cat in transactions:
        fmt_amount = f"+{a:,.2f}" if a >= 0 else f"{a:,.2f}"
        table_data.append([d, m, fmt_amount, c, _get_account_short(title)])

    col_widths = [2.4*cm, 7.5*cm, 2.6*cm, 2.0*cm, 3.0*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0),  colors.HexColor("#003366")),
        ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0),  9),
        ("ALIGN",        (0,0), (-1,0),  "CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#EEF4FF")]),
        ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",     (0,1), (-1,-1), 8),
        ("ALIGN",        (2,1), (2,-1),  "RIGHT"),
        ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ]))

    for row_idx, (_, _, amount, _, _) in enumerate(transactions, 1):
        color = colors.HexColor("#1F7A1F") if amount >= 0 else colors.HexColor("#CC0000")
        t.setStyle(TableStyle([("TEXTCOLOR", (2, row_idx), (2, row_idx), color)]))

    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    total   = sum(a for _, _, a, _, _ in transactions)
    credits = sum(a for _, _, a, _, _ in transactions if a > 0)
    debits  = sum(a for _, _, a, _, _ in transactions if a < 0)
    currency = transactions[0][3]

    summary_data = [
        ["Total Credits:", f"+{credits:,.2f} {currency}"],
        ["Total Debits:",  f"{debits:,.2f} {currency}"],
        ["Net Balance:",   f"{total:,.2f} {currency}"],
    ]
    st = Table(summary_data, colWidths=[4*cm, 5*cm])
    st.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0), (-1,-1), 9),
        ("FONTNAME",  (0,0), (0,-1),  "Helvetica-Bold"),
        ("ALIGN",     (1,0), (1,-1),  "RIGHT"),
        ("LINEABOVE", (0,0), (-1,0),  0.5, colors.HexColor("#003366")),
        ("TOPPADDING",(0,0), (-1,-1), 3),
    ]))
    story.append(st)
    doc.build(story)
    logger.info("Generated %s", path)


def _get_account_short(title):
    if "Wise" in title:
        return "Wise USD ****9134"
    return "VISA ****2847"


def generate_wise_pdf(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    _build_pdf(
        path=f"{out}/wise_usd_account_{month}.pdf",
        title="Wise – USD Account Statement",
        subtitle=f"Account Number: ****9134  |  Statement Period: {period_label(start, end)}",
        account_info=[
            "Account Holder: Juan Carlos Rodriguez",
            "Account Type: USD Borderless Account",
            "IBAN: BE30 9670 3766 9134",
        ],
        transactions=transactions,
        currency_note="All amounts in USD. This statement is generated automatically by Wise Europe SA.",
    )


def generate_visa_pdf(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    _build_pdf(
        path=f"{out}/visa_credit_card_{month}.pdf",
        title="VISA Credit Card – Monthly Statement",
        subtitle=f"Card Number: **** **** **** 2847  |  Period: {period_label(start, end)}",
        account_info=[
            "Cardholder: JUAN CARLOS RODRIGUEZ",
            "Issuer: Banco Itaú Uruguay S.A. – Tarjetas de Crédito",
            "Credit Limit: UYU 80,000  |  Available Credit: UYU 21,841",
        ],
        transactions=transactions,
        currency_note="All amounts in UYU (Uruguayan Peso). Minimum payment due: UYU 5,000 by 15th of next month.",
    )


# ── Playwright image generators ──────────────────────────────────────────────

def _statement_html(title: str, subtitle: str, account_info: list,
                    transactions: list, currency_note: str, accent: str = "#003366") -> str:
    """Render a full bank statement as self-contained HTML."""
    rows_html = ""
    for i, (d, m, a, c, cat) in enumerate(transactions):
        fmt = f"+{a:,.2f}" if a >= 0 else f"{a:,.2f}"
        color = "#1F7A1F" if a >= 0 else "#CC0000"
        bg = "#f5f8ff" if i % 2 == 0 else "#ffffff"
        rows_html += f"""
        <tr style="background:{bg}">
          <td>{d}</td>
          <td>{m}</td>
          <td style="text-align:right;color:{color};font-weight:600">{fmt}</td>
          <td style="text-align:center">{c}</td>
          <td style="color:#666">{cat}</td>
        </tr>"""

    total   = sum(a for _, _, a, _, _ in transactions)
    credits = sum(a for _, _, a, _, _ in transactions if a > 0)
    debits  = sum(a for _, _, a, _, _ in transactions if a < 0)
    currency = transactions[0][3]

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    background: #f0f2f5;
    padding: 32px;
  }}
  .card {{
    background: #fff;
    border-radius: 12px;
    box-shadow: 0 2px 16px rgba(0,0,0,.10);
    overflow: hidden;
    max-width: 900px;
    margin: 0 auto;
  }}
  .header {{
    background: {accent};
    color: #fff;
    padding: 28px 32px 22px;
  }}
  .header h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 6px; }}
  .header .sub {{ font-size: 13px; opacity: .85; }}
  .header .meta {{ font-size: 12px; opacity: .70; margin-top: 4px; }}
  .note {{
    background: #fffbe6;
    border-left: 4px solid #f0b429;
    font-size: 11px;
    color: #555;
    padding: 8px 32px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }}
  thead tr {{
    background: {accent};
    color: #fff;
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: .04em;
  }}
  thead th {{ padding: 11px 14px; text-align: left; font-weight: 600; }}
  tbody td {{ padding: 9px 14px; border-bottom: 1px solid #eef0f4; }}
  .summary {{
    display: flex;
    gap: 24px;
    padding: 18px 32px;
    background: #f8f9fc;
    border-top: 2px solid {accent};
    font-size: 13px;
  }}
  .summary .item {{ display: flex; flex-direction: column; }}
  .summary .label {{ font-size: 11px; color: #888; text-transform: uppercase; letter-spacing:.03em; }}
  .summary .val {{ font-size: 16px; font-weight: 700; margin-top: 3px; }}
  .summary .val.green {{ color: #1F7A1F; }}
  .summary .val.red   {{ color: #CC0000; }}
  .summary .val.blue  {{ color: {accent}; }}
</style>
</head>
<body>
<div class="card">
  <div class="header">
    <h1>{title}</h1>
    <div class="sub">{subtitle}</div>
    <div class="meta">{" &nbsp;·&nbsp; ".join(account_info)}</div>
  </div>
  <div class="note">{currency_note}</div>
  <table>
    <thead>
      <tr>
        <th>Date</th><th>Description</th>
        <th style="text-align:right">Amount</th>
        <th style="text-align:center">Currency</th>
        <th>Category</th>
      </tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  <div class="summary">
    <div class="item"><span class="label">Total Credits</span>
      <span class="val green">+{credits:,.2f} {currency}</span></div>
    <div class="item"><span class="label">Total Debits</span>
      <span class="val red">{debits:,.2f} {currency}</span></div>
    <div class="item"><span class="label">Net Balance</span>
      <span class="val blue">{total:,.2f} {currency}</span></div>
  </div>
</div>
</body>
</html>"""


def _screenshot_html(html: str, path: str):
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(viewport={"width": 960, "height": 800})
        page.set_content(html, wait_until="networkidle")
        page.locator(".card").screenshot(path=path)
        browser.close()


def generate_itau_image(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    path = f"{out}/itau_cuenta_corriente_{month}.png"
    html = _statement_html(
        title="Banco Itaú Uruguay – Cuenta Corriente",
        subtitle=f"Cuenta N° ****-4821  |  Período: {period_label(start, end)}",
        account_info=["Titular: JUAN CARLOS RODRIGUEZ", "Moneda: UYU"],
        transactions=transactions,
        currency_note="Extracto generado automáticamente. Sujeto a verificación.",
        accent="#FF6600",
    )
    _screenshot_html(html, path)
    logger.info("Generated %s", path)


def generate_brou_image(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    path = f"{out}/brou_caja_ahorros_{month}.png"
    html = _statement_html(
        title="BROU – Caja de Ahorros",
        subtitle=f"Cuenta N° ****-7203  |  Período: {period_label(start, end)}",
        account_info=["Titular: JUAN CARLOS RODRIGUEZ", "Moneda: UYU"],
        transactions=transactions,
        currency_note="Extracto generado automáticamente por BROU. Para consultas comuníquese con su sucursal.",
        accent="#006633",
    )
    _screenshot_html(html, path)
    logger.info("Generated %s", path)


def generate_wise_image(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    path = f"{out}/wise_usd_account_{month}.png"
    html = _statement_html(
        title="Wise – USD Account Statement",
        subtitle=f"Account ****9134  |  {period_label(start, end)}",
        account_info=["Juan Carlos Rodriguez", "USD Borderless Account", "IBAN: BE30 9670 3766 9134"],
        transactions=transactions,
        currency_note="All amounts in USD. This statement is generated automatically by Wise Europe SA.",
        accent="#163300",
    )
    _screenshot_html(html, path)
    logger.info("Generated %s", path)


def generate_visa_image(transactions, out, start, end):
    month = start.strftime("%B_%Y").lower()
    path = f"{out}/visa_credit_card_{month}.png"
    html = _statement_html(
        title="VISA Credit Card – Monthly Statement",
        subtitle=f"Card **** **** **** 2847  |  {period_label(start, end)}",
        account_info=["JUAN CARLOS RODRIGUEZ", "Issuer: Banco Itaú Uruguay S.A."],
        transactions=transactions,
        currency_note="All amounts in UYU. Minimum payment due by 15th of next month.",
        accent="#1A1F71",
    )
    _screenshot_html(html, path)
    logger.info("Generated %s", path)


# ── Receipt data ─────────────────────────────────────────────────────────────

# Categories that can produce a physical/digital receipt
RECEIPT_CATEGORIES = {
    "Groceries", "Dining", "Shopping", "Healthcare",
    "Entertainment", "Transport", "Education", "Subscriptions",
}

# Merchant registry: display name, address, phone, tax id, receipt prefix
MERCHANT_REGISTRY = {
    "SUPERMERCADO DISCO": {
        "display": "Supermercado Disco",
        "address": "Av. 18 de Julio 1224, Montevideo",
        "phone": "2400-5511",
        "rut": "21-234567-0001",
        "prefix": "DSC",
        "items": [
            ("Lácteos y refrigerados", 0.30),
            ("Panadería y pastelería",  0.15),
            ("Carnes y embutidos",      0.25),
            ("Bebidas y agua",          0.10),
            ("Limpieza del hogar",      0.20),
        ],
    },
    "SUPERMERCADO TIENDA INGLESA": {
        "display": "Tienda Inglesa",
        "address": "Luis A. de Herrera 1290, Montevideo",
        "phone": "2623-9800",
        "rut": "21-345678-0001",
        "prefix": "TIE",
        "items": [
            ("Frutas y verduras",       0.20),
            ("Carnes y pescados",       0.30),
            ("Abarrotes y conservas",   0.25),
            ("Higiene personal",        0.15),
            ("Vinos y licores",         0.10),
        ],
    },
    "FARMACIA DEL PUEBLO": {
        "display": "Farmacia del Pueblo",
        "address": "Colonia 1470, Montevideo",
        "phone": "2901-3322",
        "rut": "21-456789-0001",
        "prefix": "FDP",
        "items": [
            ("Medicamentos recetados",  0.50),
            ("Vitaminas y suplementos", 0.25),
            ("Cuidado personal",        0.25),
        ],
    },
    "FARMACIA URUGUAYA": {
        "display": "Farmacia Uruguaya",
        "address": "Bulevar Artigas 1044, Montevideo",
        "phone": "2709-4400",
        "rut": "21-567890-0001",
        "prefix": "FUY",
        "items": [
            ("Medicamentos OTC",        0.40),
            ("Cosméticos",              0.30),
            ("Insumos médicos",         0.30),
        ],
    },
    "RESTAURANT LA PASCUALA": {
        "display": "Restaurant La Pascuala",
        "address": "Ciudad Vieja, Pérez Castellano 1579, MVD",
        "phone": "2915-6672",
        "rut": "21-678901-0001",
        "prefix": "PAS",
        "items": [
            ("Plato principal (x2)",    0.55),
            ("Bebidas",                 0.20),
            ("Postre",                  0.15),
            ("Propina sugerida",        0.10),
        ],
    },
    "RESTAURANT EL PALENQUE": {
        "display": "El Palenque",
        "address": "Mercado del Puerto, Montevideo",
        "phone": "2916-4441",
        "rut": "21-789012-0001",
        "prefix": "PAL",
        "items": [
            ("Asado mixto (x2)",        0.60),
            ("Ensaladas",               0.15),
            ("Vino de la casa",         0.15),
            ("Café",                    0.10),
        ],
    },
    "RESTAURANT LA CIGALE": {
        "display": "La Cigale",
        "address": "Yi 1491, Montevideo",
        "phone": "2901-2344",
        "rut": "21-890123-0001",
        "prefix": "CIG",
        "items": [
            ("Menú degustación (x2)",   0.65),
            ("Maridaje de vinos",       0.25),
            ("Café y petit-fours",      0.10),
        ],
    },
    "RAPPI DELIVERY": {
        "display": "Rappi Delivery",
        "address": "Plataforma digital – Montevideo",
        "phone": "0800-RAPPI",
        "rut": "21-901234-0001",
        "prefix": "RPP",
        "items": [
            ("Pedido de comida",        0.80),
            ("Cargo de servicio",       0.12),
            ("Propina al repartidor",   0.08),
        ],
    },
    "YPF COMBUSTIBLES": {
        "display": "YPF Uruguay",
        "address": "Av. Italia 3900, Montevideo",
        "phone": "2600-1122",
        "rut": "21-012345-0001",
        "prefix": "YPF",
        "items": [
            ("Nafta Super 95 (litros)", 1.00),
        ],
    },
    "UBER": {
        "display": "Uber",
        "address": "Plataforma digital – Uruguay",
        "phone": "soporte@uber.com",
        "rut": "N/A (no residente)",
        "prefix": "UBR",
        "items": [
            ("Viaje UberX",             0.85),
            ("Recargo nocturno",        0.15),
        ],
    },
    "NETFLIX": {
        "display": "Netflix",
        "address": "Netflix International B.V. – online",
        "phone": "help.netflix.com",
        "rut": "N/A",
        "prefix": "NFX",
        "items": [("Suscripción mensual Standard", 1.00)],
    },
    "SPOTIFY": {
        "display": "Spotify",
        "address": "Spotify AB – online",
        "phone": "support.spotify.com",
        "rut": "N/A",
        "prefix": "SPT",
        "items": [("Suscripción Premium", 1.00)],
    },
    "CINE LIFE ALFAVILLE": {
        "display": "Cine Life Alfaville",
        "address": "Montevideo Shopping, L3-P2",
        "phone": "2628-4444",
        "rut": "21-223344-0001",
        "prefix": "CNE",
        "items": [
            ("Entrada x2",              0.70),
            ("Pochoclos y bebidas",      0.30),
        ],
    },
    "CINES LIFE ALFAVILLE": {
        "display": "Cine Life Alfaville",
        "address": "Montevideo Shopping, L3-P2",
        "phone": "2628-4444",
        "rut": "21-223344-0001",
        "prefix": "CNE",
        "items": [
            ("Entrada x2",              0.70),
            ("Pochoclos y bebidas",      0.30),
        ],
    },
    "GIMNASIO SMART FIT": {
        "display": "Smart Fit Uruguay",
        "address": "Av. 8 de Octubre 2831, Montevideo",
        "phone": "2200-7788",
        "rut": "21-334455-0001",
        "prefix": "SMF",
        "items": [("Mensualidad plan Black", 1.00)],
    },
    "LIBRERIA PAPELERIA DON": {
        "display": "Librería Don",
        "address": "Tristán Narvaja 1546, Montevideo",
        "phone": "2400-8899",
        "rut": "21-445566-0001",
        "prefix": "LBD",
        "items": [
            ("Material de escritorio",  0.50),
            ("Libros técnicos",         0.50),
        ],
    },
    "RESTAURANT PANORAMICO": {
        "display": "Panorámico Restaurant",
        "address": "Peatonal Sarandí 524, Montevideo",
        "phone": "2915-8800",
        "rut": "21-556677-0001",
        "prefix": "PAN",
        "items": [
            ("Entrada y plato (x2)",    0.60),
            ("Bebidas premium",         0.25),
            ("Postre",                  0.15),
        ],
    },
    "ZARA PUNTA CARRETAS": {
        "display": "Zara – Punta Carretas",
        "address": "Ellauri 350, Punta Carretas Shopping",
        "phone": "2711-5533",
        "rut": "21-667788-0001",
        "prefix": "ZAR",
        "items": [
            ("Prendas de vestir",       0.70),
            ("Accesorios",              0.30),
        ],
    },
    "APPLE STORE": {
        "display": "Apple Store Uruguay",
        "address": "World Trade Center, Torre 4",
        "phone": "2623-0011",
        "rut": "21-778899-0001",
        "prefix": "APL",
        "items": [("Producto Apple", 1.00)],
    },
    "DECATHLON MONTEVIDEO": {
        "display": "Decathlon Montevideo",
        "address": "Av. Luis Batlle Berres 6623",
        "phone": "2514-6000",
        "rut": "21-889900-0001",
        "prefix": "DCT",
        "items": [
            ("Indumentaria deportiva",  0.50),
            ("Equipamiento",            0.35),
            ("Accesorios",              0.15),
        ],
    },
    "HOTEL COTTAGE COLONIA": {
        "display": "Hotel Cottage Colonia",
        "address": "General Flores 172, Colonia del Sacramento",
        "phone": "4522-2189",
        "rut": "21-990011-0001",
        "prefix": "HCT",
        "items": [
            ("Habitación doble (1 noche)", 0.85),
            ("Desayuno continental",     0.15),
        ],
    },
    "LIBRERIAS EL CLUB": {
        "display": "Librerías El Club",
        "address": "18 de Julio 1230, Montevideo",
        "phone": "2902-5566",
        "rut": "21-100100-0001",
        "prefix": "LEC",
        "items": [
            ("Libros",                  0.75),
            ("Papelería",               0.25),
        ],
    },
    "RESTAURANT DON PEPERONE": {
        "display": "Don Peperone",
        "address": "Bulevar España 2900, Montevideo",
        "phone": "2708-3311",
        "rut": "21-200200-0001",
        "prefix": "DPP",
        "items": [
            ("Pizza artesanal (x2)",    0.60),
            ("Bebidas",                 0.25),
            ("Postre",                  0.15),
        ],
    },
    "RESTAURANT LA TABLA": {
        "display": "La Tabla",
        "address": "Rambla Armenia 3498, Montevideo",
        "phone": "2622-7744",
        "rut": "21-300300-0001",
        "prefix": "LTB",
        "items": [
            ("Tabla de fiambres (x2)",  0.50),
            ("Plato principal",         0.35),
            ("Vinos",                   0.15),
        ],
    },
    "IKEA ONLINE": {
        "display": "IKEA Online Store",
        "address": "ikea.com/uy – entrega a domicilio",
        "phone": "0800-IKEA",
        "rut": "N/A",
        "prefix": "IKE",
        "items": [
            ("Muebles y decoración",    0.80),
            ("Envío a domicilio",       0.20),
        ],
    },
    # Wise / digital
    "GITHUB COPILOT": {
        "display": "GitHub Copilot",
        "address": "github.com – online",
        "phone": "support.github.com",
        "rut": "N/A",
        "prefix": "GHC",
        "items": [("Copilot Individual (monthly)", 1.00)],
    },
    "CHATGPT PLUS": {
        "display": "ChatGPT Plus",
        "address": "openai.com – online",
        "phone": "help.openai.com",
        "rut": "N/A",
        "prefix": "CGP",
        "items": [("ChatGPT Plus subscription", 1.00)],
    },
    "DIGITAL OCEAN DROPLET": {
        "display": "DigitalOcean",
        "address": "digitalocean.com – online",
        "phone": "support.digitalocean.com",
        "rut": "N/A",
        "prefix": "DOC",
        "items": [("Droplet – Basic 2GB/1vCPU", 1.00)],
    },
    "ADOBE CREATIVE CLOUD": {
        "display": "Adobe Creative Cloud",
        "address": "adobe.com – online",
        "phone": "adobe.com/support",
        "rut": "N/A",
        "prefix": "ADO",
        "items": [("Creative Cloud All Apps", 1.00)],
    },
    "UDEMY COURSE PURCHASE": {
        "display": "Udemy",
        "address": "udemy.com – online",
        "phone": "support.udemy.com",
        "rut": "N/A",
        "prefix": "UDM",
        "items": [("Online course", 1.00)],
    },
    "FIGMA PROFESSIONAL": {
        "display": "Figma",
        "address": "figma.com – online",
        "phone": "help.figma.com",
        "rut": "N/A",
        "prefix": "FIG",
        "items": [("Figma Professional plan", 1.00)],
    },
    "NOTION TEAM PLAN": {
        "display": "Notion",
        "address": "notion.so – online",
        "phone": "notion.so/help",
        "rut": "N/A",
        "prefix": "NTN",
        "items": [("Notion Team Plan (per seat)", 1.00)],
    },
    "ANTHROPIC API USAGE": {
        "display": "Anthropic API",
        "address": "anthropic.com – online",
        "phone": "support.anthropic.com",
        "rut": "N/A",
        "prefix": "ANT",
        "items": [("API usage – Claude models", 1.00)],
    },
    "AMAZON WEB SERVICES": {
        "display": "Amazon Web Services",
        "address": "aws.amazon.com – online",
        "phone": "aws.amazon.com/support",
        "rut": "N/A",
        "prefix": "AWS",
        "items": [
            ("EC2 – t3.micro",          0.35),
            ("S3 Storage",              0.25),
            ("Lambda invocations",      0.20),
            ("Data transfer",           0.20),
        ],
    },
    "ZOOM SUBSCRIPTION": {
        "display": "Zoom",
        "address": "zoom.us – online",
        "phone": "support.zoom.us",
        "rut": "N/A",
        "prefix": "ZOM",
        "items": [("Zoom Pro (monthly)", 1.00)],
    },
    "1PASSWORD FAMILY": {
        "display": "1Password",
        "address": "1password.com – online",
        "phone": "support.1password.com",
        "rut": "N/A",
        "prefix": "1PW",
        "items": [("1Password Families plan", 1.00)],
    },
    "TAILSCALE VPN": {
        "display": "Tailscale",
        "address": "tailscale.com – online",
        "phone": "tailscale.com/support",
        "rut": "N/A",
        "prefix": "TLS",
        "items": [("Tailscale Personal Pro", 1.00)],
    },
    "NAMECHEAP DOMAIN": {
        "display": "Namecheap",
        "address": "namecheap.com – online",
        "phone": "support.namecheap.com",
        "rut": "N/A",
        "prefix": "NMC",
        "items": [("Domain registration / renewal", 1.00)],
    },
}


def _split_items(merchant_key: str, total: float) -> list:
    """Split total amount into line items based on merchant proportions."""
    info = MERCHANT_REGISTRY.get(merchant_key)
    if not info:
        return [("Compra / servicio", abs(total))]

    proportions = info["items"]
    items = []
    remaining = abs(total)
    for idx, (name, ratio) in enumerate(proportions):
        if idx == len(proportions) - 1:
            items.append((name, round(remaining, 2)))
        else:
            amount = round(abs(total) * ratio, 2)
            remaining -= amount
            items.append((name, amount))
    return items


def _receipt_number(prefix: str, date_str: str, seq: int) -> str:
    d = date_str.replace("-", "")
    return f"{prefix}-{d}-{seq:04d}"


# ── Receipt – PDF ────────────────────────────────────────────────────────────

def _draw_receipt_pdf(c_obj, x, y, width, merchant_key: str,
                      date_str: str, amount: float, currency: str,
                      receipt_no: str):
    """Draw one thermal-style receipt onto a reportlab canvas at (x, y)."""
    info = MERCHANT_REGISTRY.get(merchant_key, {
        "display": merchant_key.title(),
        "address": "Montevideo, Uruguay",
        "phone": "–",
        "rut": "–",
        "prefix": "RCP",
        "items": [("Compra", 1.00)],
    })

    from reportlab.lib import colors as _col

    LINE_H = 14
    cur_y  = y

    def line(text, bold=False, center=False, size=8, color=None):
        nonlocal cur_y
        c_obj.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        if color:
            c_obj.setFillColor(color)
        if center:
            c_obj.drawCentredString(x + width / 2, cur_y, text)
        else:
            c_obj.drawString(x + 8, cur_y, text)
        cur_y -= LINE_H
        c_obj.setFillColor(_col.black)

    def divider(dash=False):
        nonlocal cur_y
        c_obj.setStrokeColor(_col.HexColor("#AAAAAA"))
        if dash:
            c_obj.setDash(2, 3)
        c_obj.line(x + 4, cur_y + 8, x + width - 4, cur_y + 8)
        c_obj.setDash()
        cur_y -= 6

    line(info["display"], bold=True, center=True, size=11)
    line(info["address"],  center=True, size=7)
    line(f"Tel: {info['phone']}  RUT: {info['rut']}", center=True, size=7)
    cur_y -= 4
    divider()

    line(f"Comprobante: {receipt_no}", size=8)
    line(f"Fecha: {date_str}   Moneda: {currency}", size=8)
    divider(dash=True)

    items = _split_items(merchant_key, amount)
    for item_name, item_amt in items:
        # wrap long names
        label = textwrap.shorten(item_name, width=22, placeholder="…")
        amt_str = f"{currency} {item_amt:>10,.2f}"
        c_obj.setFont("Helvetica", 8)
        c_obj.drawString(x + 8, cur_y, label)
        c_obj.drawRightString(x + width - 8, cur_y, amt_str)
        cur_y -= LINE_H

    divider()

    # IVA breakdown (22% included)
    iva_rate = 0.22
    base = round(abs(amount) / (1 + iva_rate), 2)
    iva  = round(abs(amount) - base, 2)

    c_obj.setFont("Helvetica", 7)
    c_obj.drawString(x + 8, cur_y, "Base imponible:")
    c_obj.drawRightString(x + width - 8, cur_y, f"{currency} {base:,.2f}")
    cur_y -= 12
    c_obj.drawString(x + 8, cur_y, "IVA 22% incluido:")
    c_obj.drawRightString(x + width - 8, cur_y, f"{currency} {iva:,.2f}")
    cur_y -= 12

    divider()
    c_obj.setFont("Helvetica-Bold", 10)
    c_obj.drawString(x + 8, cur_y, "TOTAL:")
    c_obj.drawRightString(x + width - 8, cur_y, f"{currency} {abs(amount):,.2f}")
    cur_y -= 18

    c_obj.setFont("Helvetica", 7)
    c_obj.setFillColor(_col.HexColor("#555555"))
    c_obj.drawCentredString(x + width / 2, cur_y, "Gracias por su compra")
    cur_y -= 10
    c_obj.drawCentredString(x + width / 2, cur_y, "Este comprobante es válido como factura")
    c_obj.setFillColor(_col.black)

    return y - cur_y  # height used


def generate_receipts_pdf(all_transactions: dict, out: str, start: date, end: date):
    """
    Generate one PDF per account, each containing all receipt-eligible transactions
    for that account as individual thermal-style receipt blocks.
    """
    RECEIPT_W = 8 * cm   # ~80 mm thermal width
    RECEIPT_H = 8 * cm   # max height per receipt (estimate)
    PAGE_W, PAGE_H = A4
    COLS = 2
    ROWS = 3
    COL_STEP = (PAGE_W - 2 * cm) / COLS
    ROW_STEP = (PAGE_H - 3 * cm) / ROWS

    month = start.strftime("%B_%Y").lower()

    for account_name, transactions in all_transactions.items():
        eligible = [
            (d, m, a, c, cat) for d, m, a, c, cat in transactions
            if cat in RECEIPT_CATEGORIES and a < 0
        ]
        if not eligible:
            continue

        path = f"{out}/receipts_{account_name}_{month}.pdf"
        c_obj = rl_canvas.Canvas(path, pagesize=A4)

        # Watermark / header
        def draw_page_header():
            c_obj.setFont("Helvetica-Bold", 9)
            c_obj.setFillColor(colors.HexColor("#AAAAAA"))
            c_obj.drawString(cm, PAGE_H - 0.7*cm,
                f"Receipts – {account_name.upper()} – {period_label(start, end)}")
            c_obj.setFillColor(colors.black)

        draw_page_header()
        slot = 0

        for seq, (d, m, a, c, cat) in enumerate(eligible, 1):
            col = slot % COLS
            row = (slot // COLS) % ROWS

            rx = cm + col * COL_STEP + (COL_STEP - RECEIPT_W) / 2
            ry = PAGE_H - 2.5*cm - row * ROW_STEP - RECEIPT_H

            # Receipt background
            c_obj.setStrokeColor(colors.HexColor("#CCCCCC"))
            c_obj.setFillColor(colors.HexColor("#FAFAFA"))
            c_obj.roundRect(rx, ry, RECEIPT_W, RECEIPT_H, 6, fill=1, stroke=1)
            c_obj.setFillColor(colors.black)

            receipt_no = _receipt_number(
                MERCHANT_REGISTRY.get(m, {}).get("prefix", "RCP"), d, seq
            )
            _draw_receipt_pdf(c_obj, rx, ry + RECEIPT_H - 18, RECEIPT_W,
                              m, d, a, c, receipt_no)

            slot += 1
            if slot % (COLS * ROWS) == 0 and seq < len(eligible):
                c_obj.showPage()
                draw_page_header()

        c_obj.save()
        logger.info("Generated %s", path)


# ── Receipt – Image (Playwright) ─────────────────────────────────────────────

def _receipt_html_block(merchant_key: str, date_str: str, amount: float,
                        currency: str, receipt_no: str,
                        low_confidence: bool = False) -> str:
    info = MERCHANT_REGISTRY.get(merchant_key, {
        "display": merchant_key.title(),
        "address": "Montevideo, Uruguay",
        "phone": "–",
        "rut": "–",
        "prefix": "RCP",
        "items": [("Compra", 1.00)],
    })

    items = _split_items(merchant_key, amount)
    items_html = "".join(
        f'<div class="item-row"><span>{name}</span>'
        f'<span>{currency} {amt:,.2f}</span></div>'
        for name, amt in items
    )

    iva_rate = 0.22
    base = round(abs(amount) / (1 + iva_rate), 2)
    iva  = round(abs(amount) - base, 2)

    receipt_class = "receipt low-confidence" if low_confidence else "receipt"
    return f"""
<div class="{receipt_class}">
  <div class="merchant">{info['display']}</div>
  <div class="sub-info">{info['address']}</div>
  <div class="sub-info">Tel: {info['phone']}</div>
  <div class="sub-info">RUT: {info['rut']}</div>
  <div class="divider"></div>
  <div class="meta-row"><span>Comprobante</span><span>{receipt_no}</span></div>
  <div class="meta-row"><span>Fecha</span><span>{date_str}</span></div>
  <div class="divider dashed"></div>
  {items_html}
  <div class="divider"></div>
  <div class="tax-row"><span>Base imponible</span><span>{currency} {base:,.2f}</span></div>
  <div class="tax-row"><span>IVA 22%</span><span>{currency} {iva:,.2f}</span></div>
  <div class="divider"></div>
  <div class="total-row"><span>TOTAL</span><span>{currency} {abs(amount):,.2f}</span></div>
  <div class="footer">Gracias por su compra</div>
</div>"""


def _receipts_page_html(blocks: list, account_name: str, period: str) -> str:
    grid = "\n".join(blocks)
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Courier New', 'Lucida Console', monospace;
    background: #e8e8e8;
    padding: 24px;
  }}
  h2 {{
    font-family: Arial, sans-serif;
    font-size: 13px;
    color: #888;
    margin-bottom: 16px;
    letter-spacing: .04em;
  }}
  .grid {{
    display: flex;
    flex-wrap: wrap;
    gap: 20px;
  }}
  .receipt {{
    background: #fff;
    border: 1px solid #ddd;
    border-radius: 4px;
    padding: 16px 14px;
    width: 240px;
    font-size: 11px;
    box-shadow: 2px 2px 6px rgba(0,0,0,.08);
  }}
  .merchant {{
    font-size: 14px;
    font-weight: bold;
    text-align: center;
    margin-bottom: 4px;
  }}
  .sub-info {{
    text-align: center;
    font-size: 9px;
    color: #666;
    line-height: 1.5;
  }}
  .divider {{
    border-top: 1px solid #bbb;
    margin: 7px 0;
  }}
  .divider.dashed {{ border-top-style: dashed; }}
  .meta-row, .item-row, .tax-row {{
    display: flex;
    justify-content: space-between;
    font-size: 10px;
    padding: 2px 0;
    color: #333;
  }}
  .total-row {{
    display: flex;
    justify-content: space-between;
    font-size: 13px;
    font-weight: bold;
    padding: 4px 0;
  }}
  .footer {{
    text-align: center;
    font-size: 9px;
    color: #888;
    margin-top: 8px;
  }}
</style>
</head>
<body>
  <h2>RECEIPTS – {account_name.upper()} – {period}</h2>
  <div class="grid">
    {grid}
  </div>
</body>
</html>"""


# CSS for low-confidence receipt (smaller font, blur, lower contrast)
_LOW_CONFIDENCE_CSS = """
  .receipt.low-confidence {
    transform: scale(0.92);
    filter: blur(0.6px) contrast(0.88);
    opacity: 0.92;
  }
  .receipt.low-confidence .merchant { font-size: 11px; }
  .receipt.low-confidence .sub-info, .receipt.low-confidence .footer { font-size: 7px; }
  .receipt.low-confidence .meta-row, .receipt.low-confidence .item-row,
  .receipt.low-confidence .tax-row { font-size: 8px; }
  .receipt.low-confidence .total-row { font-size: 10px; }
"""


def _single_receipt_page_html(block: str, period: str, low_confidence: bool) -> str:
    """One receipt per page; optional low-confidence styling (harder to read)."""
    extra_css = _LOW_CONFIDENCE_CSS if low_confidence else ""
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Courier New', 'Lucida Console', monospace;
    background: #e8e8e8;
    padding: 24px;
    display: flex;
    justify-content: center;
    align-items: center;
    min-height: 100vh;
  }}
  .receipt {{
    background: #fff;
    border: 1px solid #ddd;
    border-radius: 4px;
    padding: 16px 14px;
    width: 240px;
    font-size: 11px;
    box-shadow: 2px 2px 6px rgba(0,0,0,.08);
  }}
  .merchant {{ font-size: 14px; font-weight: bold; text-align: center; margin-bottom: 4px; }}
  .sub-info {{ text-align: center; font-size: 9px; color: #666; line-height: 1.5; }}
  .divider {{ border-top: 1px solid #bbb; margin: 7px 0; }}
  .divider.dashed {{ border-top-style: dashed; }}
  .meta-row, .item-row, .tax-row {{ display: flex; justify-content: space-between; font-size: 10px; padding: 2px 0; color: #333; }}
  .total-row {{ display: flex; justify-content: space-between; font-size: 13px; font-weight: bold; padding: 4px 0; }}
  .footer {{ text-align: center; font-size: 9px; color: #888; margin-top: 8px; }}
  {extra_css}
</style>
</head>
<body>
  <div class="receipt-wrap">{block}</div>
</body>
</html>"""


def generate_receipts_images(all_transactions: dict, out: str, start: date, end: date):
    """Generate 5 PNGs: 3 with one clear receipt each, 2 with one low-confidence (degraded) receipt each."""
    from playwright.sync_api import sync_playwright

    month = start.strftime("%B_%Y").lower()
    period = period_label(start, end)

    # Flatten receipt-eligible from all accounts (keep account for receipt_no context)
    eligible: list[tuple[str, str, float, str, str, int]] = []
    seq = 0
    for account_name, transactions in all_transactions.items():
        for d, m, a, c, cat in transactions:
            if cat in RECEIPT_CATEGORIES and a < 0:
                seq += 1
                eligible.append((d, m, a, c, cat, seq))
        if len(eligible) >= 5:
            break
    eligible = eligible[:5]

    if len(eligible) < 5:
        logger.warning(
            "Only %d receipt-eligible transactions; generating %d receipt images.",
            len(eligible), len(eligible),
        )

    with sync_playwright() as p:
        browser = p.chromium.launch()
        for i, (d, m, a, c, cat, seq) in enumerate(eligible):
            receipt_no = _receipt_number(
                MERCHANT_REGISTRY.get(m, {}).get("prefix", "RCP"), d, seq
            )
            low_conf = i >= 3  # 4th and 5th are low-confidence
            block = _receipt_html_block(m, d, a, c, receipt_no, low_confidence=low_conf)
            html = _single_receipt_page_html(block, period, low_conf)
            path = f"{out}/receipt_{i + 1}_{month}.png"

            page = browser.new_page(viewport={"width": 400, "height": 500})
            page.set_content(html, wait_until="networkidle")
            page.screenshot(path=path)
            page.close()

            logger.info("Generated %s%s", path, " (low-confidence)" if low_conf else "")
        browser.close()


# ── Eval labels ──────────────────────────────────────────────────────────────

def _normalization_entries(
    month: str, accounts: dict[str, list[tuple]]
) -> list[dict]:
    """Expected normalized rows per generated artifact (basenames match generate_* outputs)."""
    entries: list[dict] = []
    for account_name, basenames_types in (
        (
            "Itaú Corriente",
            (
                (f"itau_cuenta_corriente_{month}.xlsx", "xlsx"),
                (f"itau_cuenta_corriente_{month}.png", "image"),
            ),
        ),
        (
            "BROU Ahorros",
            (
                (f"brou_caja_ahorros_{month}.xlsx", "xlsx"),
                (f"brou_caja_ahorros_{month}.png", "image"),
            ),
        ),
        (
            "Wise USD",
            (
                (f"wise_usd_account_{month}.pdf", "pdf"),
                (f"wise_usd_account_{month}.png", "image"),
            ),
        ),
        (
            "VISA Credit",
            (
                (f"visa_credit_card_{month}.pdf", "pdf"),
                (f"visa_credit_card_{month}.png", "image"),
            ),
        ),
    ):
        txns = accounts[account_name]
        expected_transactions = [
            {
                "date":     d,
                "merchant": m,
                "amount":   amount,
                "currency": currency,
                "account":  account_name,
            }
            for d, m, amount, currency, _cat in txns
        ]
        for basename, file_type in basenames_types:
            entries.append({
                "source_file":           basename,
                "file_type":             file_type,
                "expected_transactions": expected_transactions,
            })
    return entries


def _eval_txn_pair_key(t_a: dict, t_b: dict) -> tuple:
    return tuple(sorted([
        (t_a["date"], t_a["merchant"], t_a["amount"], t_a["account"]),
        (t_b["date"], t_b["merchant"], t_b["amount"], t_b["account"]),
    ]))


def _eval_find_txn(
    all_txns: list[dict], account: str, merchant: str, amount: float
) -> dict | None:
    for t in all_txns:
        if (
            t["account"] == account
            and t["merchant"] == merchant
            and abs(float(t["amount"]) - amount) < 0.005
        ):
            return t
    return None


def _eval_find_all_txns(
    all_txns: list[dict], account: str, merchant: str, amount: float
) -> list[dict]:
    """All rows matching account/merchant/amount (order preserved; used for rapid-fire)."""
    return [
        dict(t)
        for t in all_txns
        if (
            t["account"] == account
            and t["merchant"] == merchant
            and abs(float(t["amount"]) - amount) < 0.005
        )
    ]


def _build_suspicious_labels(all_txns: list[dict]) -> dict[str, list]:
    """Ground-truth suspicious patterns for eval (Itaú-only synthetic rows)."""
    acct = "Itaú Corriente"

    def one(merchant: str, amount: float) -> dict | None:
        t = _eval_find_txn(all_txns, acct, merchant, amount)
        return dict(t) if t else None

    should_flag: list[dict] = []

    outlier = one("RESTAURANT LA PARRILLA", -18500.0)
    ctx_a = one("EVAL DINE BASELINE 1", -1500.0)
    ctx_b = one("EVAL DINE BASELINE 2", -1600.0)
    if outlier and ctx_a and ctx_b:
        should_flag.append({
            "pattern": "outlier_amount",
            "transactions": [outlier],
            "context_transactions": [ctx_a, ctx_b],
            "reason": "Amount is clearly elevated vs same-day dining on this account",
        })

    rapid = _eval_find_all_txns(all_txns, acct, "TIENDA XYZ EVAL RAPID", -800.0)
    if len(rapid) >= 5:
        should_flag.append({
            "pattern": "rapid_fire",
            "transactions": rapid[:5],
            "context_transactions": [],
            "reason": "Multiple identical charges to the same merchant on the same day",
        })

    rnd = one("TRANSFERENCIA EVAL ROUND", -50000.0)
    if rnd:
        should_flag.append({
            "pattern": "round_number",
            "transactions": [rnd],
            "context_transactions": [],
            "reason": "Exact round amount on a transfer line",
        })

    should_not_flag: list[dict] = []
    for merchant, amount, note in (
        ("NETFLIX", -720.0, "routine subscription"),
        ("SUPERMERCADO DISCO", -2840.0, "routine grocery"),
        ("SALARIO ENERO", 52000.0, "salary income"),
        ("EVAL ALIAS NETFLIX", -720.0, "eval subscription alias"),
    ):
        t = one(merchant, amount)
        if t:
            should_not_flag.append({**t, "note": note})

    return {"should_flag": should_flag, "should_not_flag": should_not_flag}


def write_eval_labels(out_dir: Path, start: date, end: date) -> None:
    """Write data/eval_labels.json (categorization, duplicates, normalization, suspicious)."""
    accounts = {
        "Itaú Corriente": shift_transactions(ITAU_TRANSACTIONS, start, end),
        "BROU Ahorros":   shift_transactions(BROU_TRANSACTIONS, start, end),
        "Wise USD":        shift_transactions(WISE_TRANSACTIONS, start, end),
        "VISA Credit":     shift_transactions(VISA_TRANSACTIONS, start, end),
    }

    month = start.strftime("%B_%Y").lower()
    normalization = _normalization_entries(month, accounts)

    # 1. Categorization labels — one entry per transaction, all accounts
    categorization = []
    for account, txns in accounts.items():
        for d, merchant, amount, currency, category in txns:
            categorization.append({
                "date":              d,
                "merchant":          merchant,
                "amount":            amount,
                "currency":          currency,
                "account":           account,
                "expected_category": category,
            })

    # 2. Flatten all transactions for cross-account pairing
    all_txns: list[dict] = []
    for account, txns in accounts.items():
        for d, merchant, amount, currency, _cat in txns:
            all_txns.append({
                "date":     d,
                "merchant": merchant,
                "amount":   amount,
                "currency": currency,
                "account":  account,
            })

    # 3. Duplicate / non-duplicate pairs (tiers: exact, alias, fuzzy_amount, temporal, bait)
    duplicate_pairs: list[dict] = []
    seen_dup_keys: set[tuple] = set()

    def _add_dup(t_a: dict, t_b: dict, tier: str) -> None:
        key = _eval_txn_pair_key(t_a, t_b)
        if key in seen_dup_keys:
            return
        seen_dup_keys.add(key)
        duplicate_pairs.append({
            "transaction_a": t_a,
            "transaction_b": t_b,
            "is_duplicate":  True,
            "tier":            tier,
        })

    for i, t_a in enumerate(all_txns):
        for t_b in all_txns[i + 1:]:
            if t_a["account"] == t_b["account"]:
                continue
            if t_a["currency"] != t_b["currency"]:
                continue
            if t_a["merchant"] != t_b["merchant"]:
                continue
            if abs(float(t_a["amount"]) - float(t_b["amount"])) >= 0.005:
                continue

            key = _eval_txn_pair_key(t_a, t_b)
            if key in seen_dup_keys:
                continue

            days_apart = abs(
                (date.fromisoformat(t_a["date"]) - date.fromisoformat(t_b["date"])).days
            )

            if days_apart <= 3:
                seen_dup_keys.add(key)
                duplicate_pairs.append({
                    "transaction_a": t_a,
                    "transaction_b": t_b,
                    "is_duplicate":  True,
                    "tier":            "exact",
                })

    for spec in EVAL_ALIAS_DUPLICATE_SPECS:
        ta = _eval_find_txn(
            all_txns, spec["account_a"], spec["merchant_a"], spec["amount"]
        )
        tb = _eval_find_txn(
            all_txns, spec["account_b"], spec["merchant_b"], spec["amount"]
        )
        if ta is None or tb is None:
            continue
        if ta["date"] != tb["date"]:
            continue
        _add_dup(ta, tb, "alias")

    for spec in EVAL_FUZZY_DUPLICATE_SPECS:
        ta = _eval_find_txn(
            all_txns, spec["account_a"], spec["merchant_a"], spec["amount_a"]
        )
        tb = _eval_find_txn(
            all_txns, spec["account_b"], spec["merchant_b"], spec["amount_b"]
        )
        if ta is None or tb is None:
            continue
        if ta["date"] != tb["date"]:
            continue
        _add_dup(ta, tb, "fuzzy_amount")

    # Non-dups: false-positive bait (same date/amount, different merchants), then temporal
    bait_pairs: list[dict] = []
    seen_nondup_keys: set[tuple] = set()

    for i, t_a in enumerate(all_txns):
        for t_b in all_txns[i + 1:]:
            if t_a["account"] == t_b["account"]:
                continue
            if t_a["currency"] != t_b["currency"]:
                continue
            if t_a["date"] != t_b["date"]:
                continue
            if abs(float(t_a["amount"]) - float(t_b["amount"])) >= 0.005:
                continue
            if t_a["merchant"] == t_b["merchant"]:
                continue
            k = _eval_txn_pair_key(t_a, t_b)
            if k in seen_nondup_keys:
                continue
            seen_nondup_keys.add(k)
            bait_pairs.append({
                "transaction_a": t_a,
                "transaction_b": t_b,
                "is_duplicate":  False,
                "tier":            "false_positive_bait",
            })

    temporal_pairs: list[dict] = []
    for i, t_a in enumerate(all_txns):
        for t_b in all_txns[i + 1:]:
            if t_a["account"] == t_b["account"]:
                continue
            if t_a["currency"] != t_b["currency"]:
                continue
            if t_a["merchant"] != t_b["merchant"]:
                continue
            if abs(float(t_a["amount"]) - float(t_b["amount"])) >= 0.005:
                continue
            days_apart = abs(
                (date.fromisoformat(t_a["date"]) - date.fromisoformat(t_b["date"])).days
            )
            if days_apart <= 3:
                continue
            k = _eval_txn_pair_key(t_a, t_b)
            if k in seen_nondup_keys:
                continue
            seen_nondup_keys.add(k)
            temporal_pairs.append({
                "transaction_a": t_a,
                "transaction_b": t_b,
                "is_duplicate":  False,
                "tier":            "temporal",
            })

    non_duplicate_pairs: list[dict] = []
    seen_order: set[tuple] = set()
    for cand in bait_pairs + temporal_pairs:
        k = _eval_txn_pair_key(cand["transaction_a"], cand["transaction_b"])
        if k in seen_order:
            continue
        seen_order.add(k)
        non_duplicate_pairs.append(cand)
        if len(non_duplicate_pairs) >= TARGET_NON_DUPLICATE_PAIR_COUNT:
            break

    suspicious = _build_suspicious_labels(all_txns)

    labels = {
        "categorization":      categorization,
        "duplicate_pairs":     duplicate_pairs,
        "non_duplicate_pairs": non_duplicate_pairs,
        "normalization":       normalization,
        "suspicious":          suspicious,
    }

    labels_path = out_dir / "eval_labels.json"
    labels_path.write_text(json.dumps(labels, indent=2))
    n_sf = sum(len(x.get("transactions") or []) for x in suspicious["should_flag"])
    logger.info(
        "Eval labels written to %s (%d categorization, %d dup pairs, %d non-dup pairs, "
        "%d normalization, %d suspicious flag txns, %d should-not-flag)",
        labels_path,
        len(categorization),
        len(duplicate_pairs),
        len(non_duplicate_pairs),
        len(normalization),
        n_sf,
        len(suspicious["should_not_flag"]),
    )


# ── CLI entry point ──────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate synthetic bank statement & receipt sample data."
    )
    parser.add_argument(
        "--start", default=BASELINE_START.isoformat(),
        metavar="YYYY-MM-DD",
        help=f"Period start date (default: {BASELINE_START})",
    )
    parser.add_argument(
        "--end", default=BASELINE_END.isoformat(),
        metavar="YYYY-MM-DD",
        help=f"Period end date   (default: {BASELINE_END})",
    )
    parser.add_argument(
        "--out", default=DEFAULT_OUT,
        help=f"Output directory  (default: {DEFAULT_OUT})",
    )
    parser.add_argument(
        "--no-xlsx",    action="store_true", help="Skip XLSX bank statements")
    parser.add_argument(
        "--no-pdf",     action="store_true", help="Skip PDF bank statements")
    parser.add_argument(
        "--no-images",  action="store_true", help="Skip image bank statements")
    parser.add_argument(
        "--no-receipts",action="store_true", help="Skip receipt generation")
    return parser.parse_args()


def main():
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    args = parse_args()

    start = date.fromisoformat(args.start)
    end   = date.fromisoformat(args.end)
    out   = args.out

    if start > end:
        raise ValueError(f"--start ({start}) must be before --end ({end})")

    Path(out).mkdir(parents=True, exist_ok=True)
    logger.info("Output directory: %s", out)
    logger.info("Period: %s → %s", start, end)

    # Shift all transaction dates to the requested range
    itau = shift_transactions(ITAU_TRANSACTIONS, start, end)
    brou = shift_transactions(BROU_TRANSACTIONS, start, end)
    wise = shift_transactions(WISE_TRANSACTIONS, start, end)
    visa = shift_transactions(VISA_TRANSACTIONS, start, end)

    # ── XLSX ──────────────────────────────────────────────────────────────────
    if not args.no_xlsx:
        logger.info("\n── XLSX bank statements ──────────────────────────")
        generate_itau_xlsx(itau, out, start, end)
        generate_brou_xlsx(brou, out, start, end)

    # ── PDF bank statements ───────────────────────────────────────────────────
    if not args.no_pdf:
        logger.info("\n── PDF bank statements ───────────────────────────")
        generate_wise_pdf(wise, out, start, end)
        generate_visa_pdf(visa, out, start, end)

    # ── Image bank statements (Playwright) ────────────────────────────────────
    if not args.no_images:
        logger.info("\n── Image bank statements (Playwright) ────────────")
        generate_itau_image(itau, out, start, end)
        generate_brou_image(brou, out, start, end)
        generate_wise_image(wise, out, start, end)
        generate_visa_image(visa, out, start, end)

    # ── Receipts (PDF + Image) ────────────────────────────────────────────────
    if not args.no_receipts:
        logger.info("\n── Receipts ──────────────────────────────────────")
        all_tx = {
            "itau": itau,
            "brou": brou,
            "wise": wise,
            "visa": visa,
        }
        generate_receipts_pdf(all_tx, out, start, end)
        generate_receipts_images(all_tx, out, start, end)

    # ── Eval labels ───────────────────────────────────────────────────────────
    logger.info("\n── Eval labels ───────────────────────────────────")
    write_eval_labels(Path(out), start, end)

    logger.info("\nDone. All files written to %s/", out)


if __name__ == "__main__":
    main()